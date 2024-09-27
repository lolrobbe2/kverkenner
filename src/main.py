import argparse
from utils.MemberData import MemberData
from utils.Bans import Bans
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font


def export_bans_to_excel(bans, export_bans):
    """Exports specified bans to an Excel file, creating separate sheets for each if no specific bans are given."""

    # Define shorthand names for bans
    ban_name_mapping = {
        "Speelvogels": "SP",
        "Piepedollen": "PP",
        "Knapen": "KN",
        "Krabbekoningen": "KR",
        "Jonghernieuwers": "JH",
        "Hernieuwers": "HN",
        "Leiding" : "LD",
        "Ondersteunend lid" : "OD"
    }

    # Check if any specific bans are given
    if not len(export_bans):
        # If no specific bans are provided, create a single workbook for all bans
        workbook = Workbook()

        for ban in bans.bans:
            print(ban)
            # Create a new sheet with the shorthand name
            sheet = workbook.create_sheet(title=ban_name_mapping.get(ban, ban))
            # Set default font for the entire sheet
            default_font = Font(name="Verdana")  # Set the default font to Verdana
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = default_font

            # Create header
            header = sheet.cell(
                row=1, column=1, value=f"{ban.upper()} (   /{len(bans.bans[ban])})"
            )
            header.alignment = Alignment(horizontal="center", vertical="center")

            # Set font for the header
            header.font = Font(
                name="Verdana", size=14, bold=True
            )  # Font set to Verdana, 14pt

            # Create border style for header
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Apply border to the header
            header.border = thin_border

            # Merge header cells to take the entire width
            number_of_columns = 8  # Number of columns for the member data
            sheet.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=number_of_columns
            )

            # Leave an empty row between the header and the table
            empty_row = 3  # Set row 3 as the first row for the table headers

            # Write the table headers with gray fill
            table_headers = [
                "Voornaam",
                "Naam",
                "Tel 1",
                "Tel 2",
                "Straat",
                "HuisN°",
                "Gemeente",
                "Dob",
            ]
            gray_fill = PatternFill(
                start_color="bfbfbf", end_color="bfbfbf", fill_type="solid"
            )

            for col_num, header in enumerate(table_headers, start=1):
                cell = sheet.cell(row=empty_row, column=col_num, value=header)
                cell.font = Font(name="Verdana", size=11, bold=False)
                cell.fill = gray_fill  # Set background color to gray
                # No border added to the table headers as per request

            # Write the member data without borders
            for row_num, member in enumerate(bans.bans[ban], start=empty_row + 1):
                voornaam = sheet.cell(row=row_num, column=1, value=member.voornaam)
                naam = sheet.cell(row=row_num, column=2, value=member.naam)
                telefoon = sheet.cell(row=row_num, column=3, value=member.telefoon)
                extra_telefoon = sheet.cell(
                    row=row_num, column=4, value=member.extra_telefoon
                )
                straat = sheet.cell(row=row_num, column=5, value=member.straat)

                # Store huisnummer as a number and right-align
                huisnummer_cell = sheet.cell(
                    row=row_num, column=6, value=member.huisnummer
                )
                huisnummer_cell.alignment = Alignment(
                    horizontal="right"
                )  # Right-align the cell

                gemeente = sheet.cell(row=row_num, column=7, value=member.gemeente)
                geboortedatum = sheet.cell(
                    row=row_num, column=8, value=member.geboortedatum
                )

                # Set font for all member data
                voornaam.font = Font(name="Verdana", size=11, bold=False)
                naam.font = Font(name="Verdana", size=11, bold=False)
                telefoon.font = Font(name="Verdana", size=11, bold=False)
                extra_telefoon.font = Font(name="Verdana", size=11, bold=False)
                straat.font = Font(name="Verdana", size=11, bold=False)
                huisnummer_cell.font = Font(name="Verdana", size=11, bold=False)
                gemeente.font = Font(name="Verdana", size=11, bold=False)
                geboortedatum.font = Font(name="Verdana", size=11, bold=False)

            # Adjust column widths based on the max length of content
            for col_num in range(1, number_of_columns + 1):
                max_length = 0
                for row in range(1, row_num + 1):  # Check each row in the column
                    cell_value = sheet.cell(row=row, column=col_num).value
                    if cell_value:
                        max_length = max(
                            max_length, len(str(cell_value))
                        )  # Get the maximum length
                # Set column width with a little extra space
                adjusted_width = max_length + 2  # Add some padding
                # Use the column index to get the correct column letter
                column_letter = chr(
                    64 + col_num
                )  # Convert number to letter (1 -> A, 2 -> B, ...)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Add additional space between "Tel 1" and "Tel 2" columns
            extra_space = 5  # Adjust this value to increase or decrease the space
            sheet.column_dimensions[
                "D"
            ].width += extra_space  # Increase width of "Tel 2"

        # Remove the default sheet created with the workbook
        if "Sheet" in workbook.sheetnames:
            std = workbook["Sheet"]
            workbook.remove(std)

        # Save the workbook with the name indicating it contains all bans
        workbook.save("bans_export.xlsx")

    else:
        # If specific bans are given, create a separate workbook for each ban
        for ban in export_bans:
            if ban in bans.bans:
                # Create a new workbook for the export
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = ban_name_mapping.get(ban,ban)
                # Set default font for the entire workbook
                default_font = Font(name="Verdana")  # Set the default font to Verdana
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.font = default_font

                # Create header
                header = sheet.cell(
                    row=1, column=1, value=f"{ban.upper()} (   /{len(bans.bans[ban])})"
                )
                header.alignment = Alignment(horizontal="center", vertical="center")

                # Set font for the header
                header.font = Font(
                    name="Verdana", size=14, bold=True
                )  # Font set to Verdana, 14pt

                # Create border style for header
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # Apply border to the header
                header.border = thin_border

                # Merge header cells to take the entire width
                number_of_columns = 8  # Number of columns for the member data
                sheet.merge_cells(
                    start_row=1, start_column=1, end_row=1, end_column=number_of_columns
                )

                # Leave an empty row between the header and the table
                empty_row = 3  # Set row 3 as the first row for the table headers

                # Write the table headers with gray fill
                table_headers = [
                    "Voornaam",
                    "Naam",
                    "Tel 1",
                    "Tel 2",
                    "Straat",
                    "HuisN°",
                    "Gemeente",
                    "Dob",
                ]
                gray_fill = PatternFill(
                    start_color="bfbfbf", end_color="bfbfbf", fill_type="solid"
                )

                for col_num, header in enumerate(table_headers, start=1):
                    cell = sheet.cell(row=empty_row, column=col_num, value=header)
                    cell.font = Font(name="Verdana", size=11, bold=False)
                    cell.fill = gray_fill  # Set background color to gray
                    # No border added to the table headers as per request

                # Write the member data without borders
                for row_num, member in enumerate(bans.bans[ban], start=empty_row + 1):
                    voornaam = sheet.cell(row=row_num, column=1, value=member.voornaam)
                    naam = sheet.cell(row=row_num, column=2, value=member.naam)
                    telefoon = sheet.cell(row=row_num, column=3, value=member.telefoon)
                    extra_telefoon = sheet.cell(
                        row=row_num, column=4, value=member.extra_telefoon
                    )
                    straat = sheet.cell(row=row_num, column=5, value=member.straat)

                    # Store huisnummer as a number and right-align
                    huisnummer_cell = sheet.cell(
                        row=row_num, column=6, value=member.huisnummer
                    )
                    huisnummer_cell.alignment = Alignment(
                        horizontal="right"
                    )  # Right-align the cell

                    gemeente = sheet.cell(row=row_num, column=7, value=member.gemeente)
                    geboortedatum = sheet.cell(
                        row=row_num, column=8, value=member.geboortedatum
                    )

                    # Set font for all member data
                    voornaam.font = Font(name="Verdana", size=11, bold=False)
                    naam.font = Font(name="Verdana", size=11, bold=False)
                    telefoon.font = Font(name="Verdana", size=11, bold=False)
                    extra_telefoon.font = Font(name="Verdana", size=11, bold=False)
                    straat.font = Font(name="Verdana", size=11, bold=False)
                    huisnummer_cell.font = Font(name="Verdana", size=11, bold=False)
                    gemeente.font = Font(name="Verdana", size=11, bold=False)
                    geboortedatum.font = Font(name="Verdana", size=11, bold=False)

                # Adjust column widths based on the max length of content
                for col_num in range(1, number_of_columns + 1):
                    max_length = 0
                    for row in range(1, row_num + 1):  # Check each row in the column
                        cell_value = sheet.cell(row=row, column=col_num).value
                        if cell_value:
                            max_length = max(
                                max_length, len(str(cell_value))
                            )  # Get the maximum length
                    # Set column width with a little extra space
                    adjusted_width = max_length + 2  # Add some padding
                    # Use the column index to get the correct column letter
                    column_letter = chr(
                        64 + col_num
                    )  # Convert number to letter (1 -> A, 2 -> B, ...)
                    sheet.column_dimensions[column_letter].width = adjusted_width 

                # Add additional space between "Tel 1" and "Tel 2" columns
                extra_space = 5  # Adjust this value to increase or decrease the space
                sheet.column_dimensions[
                    "D"
                ].width += extra_space  # Increase width of "Tel 2"

                # Save the workbook with the name corresponding to the ban

                workbook.save(f"{ban.lower()}.xlsx")  # Save as lowercase ban name


def main(filepath, bans_to_dump, export_bans):
    # Load the Excel file
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active

    # Assuming the data starts from the second row, with the first row as headers
    members_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        member = MemberData(row=row)
        members_list.append(member)

    # Create a Bans object from the list of members
    bans = Bans(members=members_list)

    # Print the specified bans
    if bans_to_dump != None and len(bans_to_dump) != 0:
        for ban in bans_to_dump:
            if ban in bans.bans:
                # Print members in the specified ban
                bans.print_table_for_ban(ban)
            else:
                print(f"Warning: Ban '{ban}' is not recognized!")
    elif bans_to_dump != None and len(bans_to_dump) == 0:
        bans.print_table()
    if export_bans != None:
        export_bans_to_excel(bans, export_bans)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process member data.")
    parser.add_argument(
        "--filepath",
        required=True,
        help="Path to the Excel file containing member data.",
    )
    parser.add_argument(
        "--dump", "-d", nargs="*", help="List of bans to print.", default=None
    )
    parser.add_argument(
        "--export", "-e", nargs="*", help="List of bans to export to separate Excel files.", default=None
    )
    args = parser.parse_args()
    main(args.filepath, args.dump,  args.export)
